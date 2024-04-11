#!/usr/bin/env python
# -*- coding:utf-8 -*-
import os
import requests
from lxml import etree
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

headers = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7',
    'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
}

# Fetch page source
url = 'https://github.com/trending'
page_text = requests.get(url=url, headers=headers).text

# Parse data
tree = etree.HTML(page_text)

# List to store each repository's details
repo_details = []

for i in range(1, 26):  # Loop from 1 to 25
    # Extract information for each article
    details = {
        'repo_names': ' '.join(tree.xpath(f'/html/body//article[{i}]/h2/a/text()')).strip(),
        'authors_names': ' '.join(tree.xpath(f'/html/body//article[{i}]/h2/a/span/text()')).strip(),
        'repo_descriptions': ' '.join(tree.xpath(f'/html/body//article[{i}]/p/text()')).strip(),
        'languages': ' '.join(tree.xpath(f'/html/body//article[{i}]/div[@class="f6 color-fg-muted mt-2"]/span/span[2]/text()')).strip() or "None",
        'total_stars': ' '.join(tree.xpath(f'/html/body//article[{i}]/div[@class="f6 color-fg-muted mt-2"]/a[1]/text()')).strip(),
        'total_forks': ' '.join(tree.xpath(f'/html/body//article[{i}]/div[@class="f6 color-fg-muted mt-2"]/a[2]/text()')).strip(),
        'built_by': [url.strip() for url in tree.xpath(f'/html/body//article[{i}]/div[@class="f6 color-fg-muted mt-2"]/span[2]//a/img/@src')]
    }
    repo_details.append(details)

# Now, repo_details is a list of dictionaries, each containing the details of a repository.
for i, repo in enumerate(repo_details, start=1):
    print(f"Repository {i}:")
    print(f"Name: {repo['repo_names']}")
    print(f"Author: {repo['authors_names']}")
    print(f"Descriptions: {repo['repo_descriptions']}")
    print(f"Languages: {repo['languages']}")
    print(f"Stars: {repo['total_stars']}")
    print(f"Forks: {repo['total_forks']}")
    print(f"Built by: {', '.join(repo['built_by'])}")
    print("------")

# Create a new Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active
ws.title = "GitHub_Trending_Data_04_11"

# Define and add the header row to the worksheet
headers = ['Repo Names', 'Authors Names', 'Descriptions', 'Languages', 'Total Stars', 'Total Forks', 'Built By URLs']
ws.append(headers)

# Iterate through the list of repository details
for repo in repo_details:
    # For each repository, add a row to the worksheet
    row = [
        repo['repo_names'],
        repo['authors_names'],
        repo['repo_descriptions'],
        repo['languages'],
        repo['total_stars'],
        repo['total_forks'],
        ', '.join(repo['built_by'])  # Join URLs with comma for Excel cell
    ]
    ws.append(row)

# Define the file path and name for the Excel file
excel_file_path = "GitHub_Trending_Data_04_11.xlsx"

# Adjust column widths based on the longest entry in each column
for column in ws.columns:
    max_length = 0
    for cell in column:
        # Ensure the cell value is a string to measure its length
        cell_length = len(str(cell.value))
        if cell_length > max_length:
            max_length = cell_length
    # Adjust the width; customize the multiplier or added as needed for aesthetics
    adjusted_width = max_length + 2
    # Get the column letter and set the dimension
    column_letter = get_column_letter(column[0].column)
    ws.column_dimensions[column_letter].width = adjusted_width

# Save the workbook to the defined file
wb.save(excel_file_path)

print(f"Data successfully saved to {excel_file_path}")


